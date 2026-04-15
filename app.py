import io
import os
import re
from datetime import datetime
from difflib import get_close_matches
from openpyxl.styles import Alignment, Font, PatternFill
import pandas as pd
import pdfplumber
from flask import Flask, request, render_template_string, send_file

# Optional OCR fallback
import pytesseract
from pdf2image import convert_from_bytes

app = Flask(__name__)
LAST_ROWS = []

# -------------------- OPTIONAL TESSERACT PATH --------------------
TESS_PATH = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
if os.path.exists(TESS_PATH):
    pytesseract.pytesseract.tesseract_cmd = TESS_PATH

# =========================================================
# COMMON HELPERS
# =========================================================
def clean_text(x):
    return "" if x is None else str(x).strip()


def to_float(x):
    try:
        if x is None:
            return 0.0
        s = str(x).strip().replace(",", "")
        return float(s) if s else 0.0
    except Exception:
        return 0.0


def normalize_name(x: str) -> str:
    s = clean_text(x).upper().replace("\n", " ").replace("\r", " ")
    s = re.sub(r"[^A-Z0-9()& /-]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def simple_key(x: str) -> str:
    s = normalize_name(x)
    return re.sub(r"[^A-Z]", "", s)


CITY_STATE_LOOKUP = {}
CITY_KEYS_SORTED = []
CITY_DATASET_STATUS = ""


def load_india_cities_csv():
    global CITY_STATE_LOOKUP, CITY_KEYS_SORTED, CITY_DATASET_STATUS
    CITY_STATE_LOOKUP = {}
    CITY_KEYS_SORTED = []
   #CITY_DATASET_STATUS = "india_cities.csv not loaded. Using built-in city mappings only."

    path = "india_cities.csv"
    if not os.path.exists(path):
        return

    try:
        df = pd.read_csv(path, low_memory=False)
        cols = {c.lower(): c for c in df.columns}

        if "country_name" in cols and "name" in cols:
            c_country = cols["country_name"]
            c_city = cols["name"]

            for _, r in df[df[c_country].astype(str).str.upper().eq("INDIA")].iterrows():
                city = simple_key(r.get(c_city, ""))
                if city:
                    CITY_STATE_LOOKUP.setdefault(city, "INDIA")

        elif "name" in cols:
            c_city = cols["name"]

            for _, r in df.iterrows():
                city = simple_key(r.get(c_city, ""))
                if city:
                    CITY_STATE_LOOKUP.setdefault(city, "INDIA")

        CITY_KEYS_SORTED = sorted(CITY_STATE_LOOKUP.keys(), key=len, reverse=True)
        CITY_DATASET_STATUS = f"india_cities.csv loaded. Cities indexed: {len(CITY_STATE_LOOKUP)}"

    except Exception as e:
        CITY_STATE_LOOKUP = {}
        CITY_KEYS_SORTED = []
        CITY_DATASET_STATUS = f"india_cities.csv load failed: {e}. Using built-in city mappings only."


def best_alias_match(name: str, mapping: dict, cutoff: float = 0.78):
    if not name:
        return None

    n = normalize_name(name)
    if n in mapping:
        return mapping[n]

    nk = simple_key(n)
    for k, v in mapping.items():
        if simple_key(k) == nk:
            return v

    keys = list(mapping.keys())
    m = get_close_matches(n, keys, n=1, cutoff=cutoff)
    if m:
        return mapping[m[0]]

    compact_map = {simple_key(k): v for k, v in mapping.items()}
    choices = list(compact_map.keys())
    m2 = get_close_matches(nk, choices, n=1, cutoff=cutoff)
    if m2:
        return compact_map[m2[0]]

    return None


def ocr_extract_text(pdf_bytes: bytes, max_pages=3):
    imgs = convert_from_bytes(pdf_bytes, dpi=220, first_page=1, last_page=max_pages)
    texts = []
    for im in imgs:
        texts.append(pytesseract.image_to_string(im))
    return "\n".join(texts)


# =========================================================
# DTDC CONFIG
# =========================================================
DTDC_SURFACE_RULES = {
   # "MIN_FREIGHT_PER_DOCKET": 400.0,
    "DOCKET_CHARGE": 100.0,
    "FSC_PERCENT": 0.20,
    "ROV_PERCENT": 0.001,
    "ROV_MIN": 100.0,
    "COD_MIN": 250.0,
}

DTDC_SURFACE_ZONE_MATRIX = {
    "North I": {"North I": 9, "North II": 10, "North III": 11, "West I": 12.75, "West II": 14.75, "Central I": 13.75, "Central II": 14.75, "South I": 14.75, "South II": 16.75, "South III": 17.75, "East I": 15.75, "East II": 16.75, "East III": 20.75},
    "North II": {"North I": 10, "North II": 10, "North III": 10, "West I": 14.75, "West II": 15.75, "Central I": 15.75, "Central II": 17.75, "South I": 16.75, "South II": 17.75, "South III": 18.75, "East I": 17.75, "East II": 18.75, "East III": 20.75},
    "North III": {"North I": 10, "North II": 10, "North III": 10, "West I": 11, "West II": 15.75, "Central I": 17.75, "Central II": 17.75, "South I": 17.75, "South II": 18.75, "South III": 20.75, "East I": 17.75, "East II": 18.75, "East III": 22.75},
    "West I": {"North I": 13.75, "North II": 14.75, "North III": 14.75, "West I": 9, "West II": 9, "Central I": 10, "Central II": 11, "South I": 12.75, "South II": 14.75, "South III": 14.75, "East I": 17.75, "East II": 17.75, "East III": 22.75},
    "West II": {"North I": 14.75, "North II": 15.75, "North III": 17.5, "West I": 9, "West II": 10, "Central I": 11, "Central II": 11, "South I": 13.75, "South II": 14.75, "South III": 14.75, "East I": 15.75, "East II": 17.75, "East III": 22.75},
    "Central I": {"North I": 13.75, "North II": 15.75, "North III": 17.5, "West I": 10, "West II": 11, "Central I": 9, "Central II": 10, "South I": 13.75, "South II": 14.75, "South III": 15.75, "East I": 16.75, "East II": 17.75, "East III": 21.75},
    "Central II": {"North I": 14.75, "North II": 17.75, "North III": 17.75, "West I": 11, "West II": 11, "Central I": 10, "Central II": 9, "South I": 14.75, "South II": 15.75, "South III": 16.75, "East I": 17.75, "East II": 17.75, "East III": 22.75},
    "South I": {"North I": 14.75, "North II": 15.75, "North III": 16.75, "West I": 12.75, "West II": 13.75, "Central I": 13.75, "Central II": 13.75, "South I": 8, "South II": 9, "South III": 10, "East I": 17.75, "East II": 18.75, "East III": 22.75},
    "South II": {"North I": 14.75, "North II": 15.75, "North III": 17.75, "West I": 13.75, "West II": 13.75, "Central I": 13.75, "Central II": 14.75, "South I": 8, "South II": 9, "South III": 10, "East I": 19.75, "East II": 19.75, "East III": 25.75},
    "South III": {"North I": 16.75, "North II": 17.75, "North III": 19.75, "West I": 14.75, "West II": 15.75, "Central I": 14.75, "Central II": 16.75, "South I": 9, "South II": 10, "South III": 10, "East I": 19.75, "East II": 19.75, "East III": 25.75},
    "East I": {"North I": 13.75, "North II": 14.75, "North III": 14.75, "West I": 14.75, "West II": 14.75, "Central I": 13.75, "Central II": 14.75, "South I": 15.75, "South II": 16.75, "South III": 16.75, "East I": 9, "East II": 10, "East III": 11.75},
    "East II": {"North I": 14.75, "North II": 14.75, "North III": 15.75, "West I": 14.75, "West II": 15.75, "Central I": 14.75, "Central II": 14.75, "South I": 15.75, "South II": 16.75, "South III": 17.75, "East I": 10, "East II": 11, "East III": 11.75},
    "East III": {"North I": 14.75, "North II": 14.75, "North III": 15.75, "West I": 16.75, "West II": 17.75, "Central I": 16.75, "Central II": 18.75, "South I": 17.75, "South II": 22.75, "South III": 24.75, "East I": 13.75, "East II": 12.75, "East III": 13.75},
}

DTDC_CITY_ALIASES = {
    "MADUR": "MADURAI",
    "MADUR AI": "MADURAI",
    "CHEN NAI": "CHENNAI",
    "BANG ALORE": "BANGALORE",
    "BANG ALOR E": "BANGALORE",
    "PANT NAGAR": "PANTNAGAR",
    "COIMB": "COIMBATORE",
    "COIMB ATORE": "COIMBATORE",
    "AHMEDABA": "AHMEDABAD",
    "AHMED ABAD": "AHMEDABAD",
    "TIRUVALLORE": "TIRUVALLUR",
    "TIRUVALLURE": "TIRUVALLUR",
    "PALGHAT": "PALAKKAD",
    "PALG HAT": "PALAKKAD",
    "COCHIN": "KOCHI",
    "COCH IN": "KOCHI",
    "COCHI N": "KOCHI",
    "NASIK": "NASHIK",
    "BARODA": "VADODARA",
}

DTDC_CITY_ZONE = {
    "DELHI": "North I",
    "FARIDABAD": "North I",
    "GHAZIABAD": "North I",
    "GURGAON": "North I",
    "GURUGRAM": "North I",
    "NOIDA": "North I",
    "KUNDLI": "North I",

    "RAJPURA": "North II",
    "CHANDIGARH": "North II",
    "PANTNAGAR": "North II",
    "MANESAR": "North II",
    "REWARI": "North II",
    "BALLABGARH": "North II",
    "ROORKEE": "North II",
    "BADDI": "North II",

    "JAMMU": "North III",
    "SHIMLA": "North III",

    "BHUBANESWAR": "East I",
    "JAMSHEDPUR": "East I",
    "KOLKATA": "East I",
    "PATNA": "East I",

    "HOOGLY": "East II",
    "GAMARIA": "East II",

    "AHMEDABAD": "West I",
    "VADODARA": "West I",
    "BARODA": "West I",
    "MUMBAI": "West I",
    "THANE": "West I",
    "PUNE": "West I",

    "SANTEJ": "West II",
    "GANDHINAGAR": "West II",
    "BARDOLI": "West II",
    "VASAI": "West II",
    "RAIGAD": "West II",
    "AHMEDNAGAR": "West II",
    "NASHIK": "West II",
    "NAVI MUMBAI": "West II",
    "SANASWADI": "West II",
    "PUNE CHAKAN": "West II",
    "SHIRUR": "West II",

    "BANGALORE": "South I",
    "BENGALURU": "South I",
    "CHENNAI": "South I",
    "HYDERABAD": "South I",
    "SECUNDERABAD": "South I",
    "SRIPERUMBUDUR": "South I",
    "BANGALORE DELIVERY": "South I",

    "MADURAI": "South II",
    "HOSUR": "South II",
    "COIMBATORE": "South II",
    "ERODE": "South II",
    "SALEM": "South II",
    "METTUPALAYAM": "South II",
    "TRICHY": "South II",
    "TIRUCHIRAPPALLI": "South II",
    "TIRUVALLUR": "South II",
    "MELUR": "South II",

    "PALAKKAD": "South III",
    "KOCHI": "South III",

    "BHOPAL": "Central I",
    "INDORE": "Central I",
    "RAIPUR": "Central I",
    "NAGPUR": "Central I",
}

# =========================================================
# SAFEXPRESS CONFIG
# =========================================================
SAFE_ORIGIN_CITY = "MADURAI"
SAFE_ORIGIN_ZONE = "SOUTH ONE"

SAFE_SURFACE_RULES = {
    "RATE_BY_MODE": {"A": 6.0, "B": 8.0, "C": 10.0, "D": 12.0, "E": 15.0},
    "MIN_FREIGHT": 500.0,
    "MIN_WEIGHT": 20.0,
    "WAYBILL_CHARGE": 150.0,
    "UCC_CHARGE": 200.0,
    "FUEL_PERCENT": 0.10,
    "VALUE_SURCHARGE_PERCENT": 0.001,
    "VALUE_SURCHARGE_MIN": 100.0,
    "GST_PERCENT": 0.18,
    "AREA_5_PER_KG": 5.0,
    "AREA_10_PER_KG": 10.0,
    "AREA_15_PER_KG": 15.0,
}

SAFE_AIR_RULES = {
    "RATE_BY_MODE": {"A": 50.0, "B": 60.0, "C": 70.0, "D": 80.0, "E": 120.0},
    "MIN_FREIGHT": 1000.0,
    "MIN_WEIGHT": 10.0,
    "WAYBILL_CHARGE": 150.0,
    "UCC_CHARGE": 200.0,
    "FUEL_PERCENT": 0.10,
    "VALUE_SURCHARGE_PERCENT": 0.001,
    "VALUE_SURCHARGE_MIN": 100.0,
    "GST_PERCENT": 0.18,
    "AREA_5_PER_KG": 5.0,
    "AREA_10_PER_KG": 10.0,
    "AREA_15_PER_KG": 15.0,
    "ISLAND_RATE": 160.0,
}

SAFE_ZONE_MATRIX = {
    "NORTH ONE": {"NORTH ONE": "A", "NORTH TWO": "A", "EAST": "D", "NORTHEAST": "E", "WEST ONE": "B", "WEST TWO": "C", "SOUTH ONE": "C", "SOUTH TWO": "D", "CENTRAL": "B"},
    "NORTH TWO": {"NORTH ONE": "A", "NORTH TWO": "A", "EAST": "D", "NORTHEAST": "E", "WEST ONE": "C", "WEST TWO": "C", "SOUTH ONE": "D", "SOUTH TWO": "D", "CENTRAL": "B"},
    "EAST": {"NORTH ONE": "C", "NORTH TWO": "D", "EAST": "A", "NORTHEAST": "B", "WEST ONE": "C", "WEST TWO": "D", "SOUTH ONE": "C", "SOUTH TWO": "D", "CENTRAL": "B"},
    "NORTHEAST": {"NORTH ONE": "C", "NORTH TWO": "D", "EAST": "B", "NORTHEAST": "A", "WEST ONE": "D", "WEST TWO": "D", "SOUTH ONE": "D", "SOUTH TWO": "B", "CENTRAL": "C"},
    "WEST ONE": {"NORTH ONE": "B", "NORTH TWO": "C", "EAST": "D", "NORTHEAST": "E", "WEST ONE": "A", "WEST TWO": "A", "SOUTH ONE": "C", "SOUTH TWO": "D", "CENTRAL": "B"},
    "WEST TWO": {"NORTH ONE": "C", "NORTH TWO": "D", "EAST": "D", "NORTHEAST": "E", "WEST ONE": "A", "WEST TWO": "A", "SOUTH ONE": "B", "SOUTH TWO": "D", "CENTRAL": "B"},
    "SOUTH ONE": {"NORTH ONE": "C", "NORTH TWO": "D", "EAST": "D", "NORTHEAST": "E", "WEST ONE": "C", "WEST TWO": "B", "SOUTH ONE": "A", "SOUTH TWO": "B", "CENTRAL": "B"},
    "SOUTH TWO": {"NORTH ONE": "D", "NORTH TWO": "D", "EAST": "D", "NORTHEAST": "E", "WEST ONE": "C", "WEST TWO": "C", "SOUTH ONE": "A", "SOUTH TWO": "A", "CENTRAL": "B"},
    "CENTRAL": {"NORTH ONE": "B", "NORTH TWO": "C", "EAST": "D", "NORTHEAST": "E", "WEST ONE": "A", "WEST TWO": "B", "SOUTH ONE": "B", "SOUTH TWO": "D", "CENTRAL": "A"},
}

SAFE_CITY_ALIASES = {
    "BANG ALORE": "BANGALORE",
    "BANG ALOR E": "BANGALORE",
    "BANGALORE DELIVERY": "BANGALORE",
    "COCH IN": "COCHIN",
    "COCHI N": "COCHIN",
    "PALG HAT": "PALGHAT",
    "PALAKKAD": "PALGHAT",
    "SHIRUR (MH)": "SHIRUR",
    "SRI PERUM BUDUR": "SRIPERUMBUDUR",
    "SRI PERUMBUDUR": "SRIPERUMBUDUR",
}

SAFE_CITY_ZONE = {
    "DELHI": "NORTH ONE",
    "NOIDA": "NORTH ONE",
    "FARIDABAD": "NORTH ONE",
    "GURGAON": "NORTH ONE",
    "GURUGRAM": "NORTH ONE",

    "CHANDIGARH": "NORTH TWO",
    "MANESAR": "NORTH TWO",
    "REWARI": "NORTH TWO",
    "ROORKEE": "NORTH TWO",
    "BADDI": "NORTH TWO",

    "KOLKATA": "EAST",
    "HOOGLY": "EAST",
    "PATNA": "EAST",
    "JAMSHEDPUR": "EAST",
    "GAMARIA": "EAST",

    "GUWAHATI": "NORTHEAST",

    "AHMEDABAD": "WEST ONE",
    "VADODARA": "WEST ONE",
    "BARODA": "WEST ONE",
    "GANDHINAGAR": "WEST ONE",
    "SANTEJ": "WEST ONE",
    "BARDOLI": "WEST ONE",

    "MUMBAI": "WEST TWO",
    "NAVI MUMBAI": "WEST TWO",
    "VASAI": "WEST TWO",
    "PUNE": "WEST TWO",
    "PUNE CHAKAN": "WEST TWO",
    "SANASWADI": "WEST TWO",
    "SHIRUR": "WEST TWO",
    "NASHIK": "WEST TWO",
    "AHMEDNAGAR": "WEST TWO",
    "GOA": "WEST TWO",

    "MADURAI": "SOUTH ONE",
    "CHENNAI": "SOUTH ONE",
    "BANGALORE": "SOUTH ONE",
    "BENGALURU": "SOUTH ONE",
    "HYDERABAD": "SOUTH ONE",
    "SECUNDERABAD": "SOUTH ONE",
    "SRIPERUMBUDUR": "SOUTH ONE",
    "COIMBATORE": "SOUTH ONE",
    "HOSUR": "SOUTH ONE",
    "ERODE": "SOUTH ONE",
    "SALEM": "SOUTH ONE",
    "TRICHY": "SOUTH ONE",
    "TIRUCHIRAPPALLI": "SOUTH ONE",

    "PALGHAT": "SOUTH TWO",
    "COCHIN": "SOUTH TWO",
    "KOCHI": "SOUTH TWO",
    "PUDUCHERRY": "SOUTH TWO",
    "PONDICHERRY": "SOUTH TWO",

    "BHOPAL": "CENTRAL",
    "INDORE": "CENTRAL",
    "RAIPUR": "CENTRAL",
    "NAGPUR": "CENTRAL",
}

SAFE_UCC_CITIES = {
    "AHMEDABAD", "BANGALORE", "BENGALURU", "CHENNAI", "DELHI",
    "HYDERABAD", "KOLKATA", "MUMBAI", "PUNE"
}

AREA_5_CITIES = {"PALGHAT", "COCHIN"}
AREA_10_CITIES = set()
AREA_15_CITIES = set()
ISLAND_CITIES = {"ANDAMAN", "LAKSHADWEEP", "LADAKH"}

# =========================================================
# GENERIC PDF TABLE EXTRACTION
# =========================================================
def header_score(row):
    t = " ".join([str(c).upper() for c in row if c is not None])
    keys = ["DOCKET", "BKG", "DLY", "CHARGED", "WEIGHT", "RATE", "FREIGHT", "FSC", "TOTAL", "WAYBILL", "WGHT", "DESTINATION"]
    return sum(1 for k in keys if k in t)


def extract_tables_from_pdf(pdf_bytes: bytes):
    dfs = []
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for page in pdf.pages:
            tables = page.extract_tables() or []
            for table in tables:
                if not table or len(table) < 2:
                    continue

                best_i, best_s = None, 0
                for i in range(min(12, len(table))):
                    s = header_score(table[i])
                    if s > best_s:
                        best_s, best_i = s, i

                if best_s < 3 or best_i is None:
                    continue

                header = [str(c).strip() if c else "" for c in table[best_i]]
                data = table[best_i + 1:]
                if not data:
                    continue
                dfs.append(pd.DataFrame(data, columns=header))
    return dfs

# =========================================================
# DTDC
# =========================================================
def normalize_dtdc_city(city_raw: str) -> str:
    c = normalize_name(city_raw)
    if c in DTDC_CITY_ALIASES:
        c = DTDC_CITY_ALIASES[c]
    return c


def resolve_dtdc_zone(city_raw: str):
    c = normalize_dtdc_city(city_raw)
    return best_alias_match(c, DTDC_CITY_ZONE, cutoff=0.75)


def parse_dtdc_rows(pdf_bytes: bytes):
    dfs = extract_tables_from_pdf(pdf_bytes)
    rows = []

    for df in dfs:
        rename = {}
        for c in df.columns:
            u = normalize_name(c)
            if "DOCKET" in u:
                rename[c] = "Docket No"
            elif ("BKG" in u or "BOOK" in u) and "CITY" in u:
                rename[c] = "Bkg City"
            elif ("DLY" in u or "DEL" in u) and "CITY" in u:
                rename[c] = "Dly City"
            elif "CHARGED" in u and "WEIGHT" in u:
                rename[c] = "Charged Weight"
            elif "RATE" in u and ("KG" in u or "/KG" in u):
                rename[c] = "Rate/KG"
            elif "FREIGHT" in u:
                rename[c] = "Freight AMT"
            elif "FSC" in u:
                rename[c] = "FSC Amount"
            elif "TOTAL" in u:
                rename[c] = "Total Charge"
            elif "ROV" in u:
                rename[c] = "ROV Charge"
            elif "COD" in u:
                rename[c] = "COD Charge"
            elif "ODA" in u:
                rename[c] = "ODA Charge"
            elif "HANDLING" in u:
                rename[c] = "Handling Charge"
            elif "MISC" in u:
                rename[c] = "Misc Charge"
            elif ("INVOICE" in u and "VALUE" in u) or ("DECLARED" in u and "VALUE" in u):
                rename[c] = "Invoice Value"
            elif ("COD" in u and "VALUE" in u) or ("DOD" in u and "VALUE" in u):
                rename[c] = "COD Value"

        df = df.rename(columns=rename)
        df = df.loc[:, ~df.columns.duplicated()]

        must = {"Docket No", "Bkg City", "Dly City", "Charged Weight", "Total Charge"}
        if not must.issubset(df.columns):
            continue

        df["Docket No"] = df["Docket No"].astype(str).str.strip()
        df = df[df["Docket No"].str.len() >= 6]
        df = df[~df["Docket No"].str.upper().isin(["SUB TOTAL", "PAGE TOTAL", "GRAND TOTAL"])]

        for _, r in df.iterrows():
            rows.append(r.to_dict())

    return rows


def audit_dtdc_surface(row: dict):
    rules = DTDC_SURFACE_RULES
    matrix = DTDC_SURFACE_ZONE_MATRIX

    docket = clean_text(row.get("Docket No", ""))
    origin_raw = clean_text(row.get("Bkg City", ""))
    dest_raw = clean_text(row.get("Dly City", ""))
    wt = to_float(row.get("Charged Weight", 0))
    inv_rate = to_float(row.get("Rate/KG", 0))
    inv_freight = to_float(row.get("Freight AMT", 0))
    inv_fsc = to_float(row.get("FSC Amount", 0))
    inv_total = to_float(row.get("Total Charge", 0))
    inv_rov_charge = to_float(row.get("ROV Charge", 0))
    inv_cod_charge = to_float(row.get("COD Charge", 0))
    inv_oda = to_float(row.get("ODA Charge", 0))
    inv_handling = to_float(row.get("Handling Charge", 0))
    inv_misc = to_float(row.get("Misc Charge", 0))
    inv_invoice_value = to_float(row.get("Invoice Value", 0))
    inv_cod_value = to_float(row.get("COD Value", 0))

    zo = resolve_dtdc_zone(origin_raw)
    zd = resolve_dtdc_zone(dest_raw)

    cor_rate = 0.0
    cor_freight = inv_freight
    cor_fsc = inv_fsc

    if zo and zd:
        try:
            cor_rate = float(matrix[zo][zd])
            cor_freight = max(wt * cor_rate, float(rules["MIN_FREIGHT_PER_DOCKET"]))
            cor_fsc = cor_freight * float(rules["FSC_PERCENT"])
        except Exception:
            pass

    cor_docket = float(rules["DOCKET_CHARGE"])

    cor_rov = 0.0
    if inv_rov_charge > 0:
        if inv_invoice_value > 0:
            cor_rov = max(inv_invoice_value * float(rules["ROV_PERCENT"]), float(rules["ROV_MIN"]))
        else:
            cor_rov = inv_rov_charge

    cor_cod = 0.0
    if inv_cod_charge > 0:
        if inv_cod_value > 0:
            cor_cod = max(float(rules["COD_MIN"]), inv_cod_value)
        else:
            cor_cod = inv_cod_charge

    cor_oda = inv_oda
    cor_handling = inv_handling
    cor_misc = 0.0

    cor_total = cor_freight + cor_fsc + cor_docket + cor_rov + cor_cod + cor_oda + cor_handling + cor_misc
    diff = inv_total - cor_total

    if abs(diff) < 1.0:
        remark = "OK"
    elif diff > 0:
        remark = "OVERCHARGED"
    else:
        remark = "UNDERCHARGED"

    breakdown = (
        f"Courier: DTDC | Mode: Surface | Zones: {zo or 'UNKNOWN'} → {zd or 'UNKNOWN'} | "
        f"Correct Rate: {cor_rate:.2f} | Freight: {cor_freight:.2f} | FSC: {cor_fsc:.2f} | "
        f"Docket: {cor_docket:.2f} | ROV: {cor_rov:.2f} | COD: {cor_cod:.2f} | "
        f"ODA: {cor_oda:.2f} | Handling: {cor_handling:.2f} | Misc: {cor_misc:.2f}"
    )

    return {
        "Docket No": docket,
        "Origin": normalize_dtdc_city(origin_raw),
        "Destination": normalize_dtdc_city(dest_raw),
        "Zone Origin": zo or "UNKNOWN",
        "Zone Dest": zd or "UNKNOWN",
        "Mode": "Surface",
        "Weight KG": round(wt, 3),
        "Invoice Amount": round(inv_total, 2),
        "Correct Amount": round(cor_total, 2),
        "Difference": round(diff, 2),
        "Remark": remark,
        "Breakdown": breakdown,
    }

# =========================================================
# SAFEXPRESS
# =========================================================
def normalize_safe_city(city_raw: str) -> str:
    c = normalize_name(city_raw)
    if c in SAFE_CITY_ALIASES:
        c = SAFE_CITY_ALIASES[c]
    c = re.sub(r"\(MH\)", "", c).strip()
    c = c.replace(" DELIVERY", "").strip()
    c = re.sub(r"\s+", " ", c).strip()
    return c


def resolve_safe_zone(city_raw: str) -> str:
    c = normalize_safe_city(city_raw)
    return SAFE_CITY_ZONE.get(c, "UNKNOWN")


def parse_safeexpress_rows(pdf_bytes: bytes):
    rows = []
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for page in pdf.pages:
            txt = page.extract_text() or ""
            lines = [ln.strip() for ln in txt.splitlines() if ln.strip()]

            for ln in lines:
                m = re.match(
                    r"^\d+\s+(\d{8,})\s+(\d{2}-\d{2}-\d{4})\s+(.+?)\s+(\d+)\s+(\d+(?:\.\d+)?)\s+(\d+(?:\.\d+)?)$",
                    ln
                )
                if not m:
                    continue

                waybill = m.group(1)
                pickup_date = m.group(2)
                destination = m.group(3).strip()
                pkgs = m.group(4)
                weight = m.group(5)
                total_amount = m.group(6)

                rows.append({
                    "Waybill No": waybill,
                    "Pickup Date": pickup_date,
                    "Origin": SAFE_ORIGIN_CITY,
                    "Destination": normalize_safe_city(destination),
                    "Pkgs": pkgs,
                    "Weight": weight,
                    "Invoice Amount": total_amount
                })
    return rows


def get_area_charge_per_kg(dest_city: str, rules: dict) -> float:
    c = normalize_safe_city(dest_city)
    if c in AREA_15_CITIES:
        return rules["AREA_15_PER_KG"]
    if c in AREA_10_CITIES:
        return rules["AREA_10_PER_KG"]
    if c in AREA_5_CITIES:
        return rules["AREA_5_PER_KG"]
    return 0.0


def get_ucc_charge(dest_city: str, rules: dict) -> float:
    c = normalize_safe_city(dest_city)
    return rules["UCC_CHARGE"] if c in SAFE_UCC_CITIES else 0.0


def calculate_safeexpress(row: dict, mode: str):
    rules = SAFE_SURFACE_RULES if mode == "surface" else SAFE_AIR_RULES

    waybill = clean_text(row.get("Waybill No", ""))
    pickup_date = clean_text(row.get("Pickup Date", ""))
    origin = normalize_safe_city(row.get("Origin", ""))
    dest = normalize_safe_city(row.get("Destination", ""))
    actual_weight = to_float(row.get("Weight", 0))
    invoice_amount = to_float(row.get("Invoice Amount", 0))

    origin_zone = SAFE_ORIGIN_ZONE
    dest_zone = resolve_safe_zone(dest)

    mode_letter = ""
    rate = 0.0

    if dest_zone != "UNKNOWN":
        mode_letter = SAFE_ZONE_MATRIX[origin_zone][dest_zone]
        rate = rules["RATE_BY_MODE"][mode_letter]

    chargeable_weight = max(actual_weight, rules["MIN_WEIGHT"])

    if mode == "air" and dest in ISLAND_CITIES:
        freight = max(chargeable_weight * rules["ISLAND_RATE"], rules["MIN_FREIGHT"])
    else:
        freight = max(chargeable_weight * rate, rules["MIN_FREIGHT"])

    fuel = freight * rules["FUEL_PERCENT"]
    waybill_charge = rules["WAYBILL_CHARGE"]
    ucc = get_ucc_charge(dest, rules)
    area_charge = get_area_charge_per_kg(dest, rules) * chargeable_weight
    value_surcharge = max(invoice_amount * 0.001, 100)

    subtotal = freight + fuel + waybill_charge + ucc + value_surcharge + area_charge
    gst = subtotal * rules["GST_PERCENT"]
    correct_total = subtotal + gst

    diff = invoice_amount - correct_total

    if abs(diff) < 1.0:
        remark = "OK"
    elif diff > 0:
        remark = "OVERCHARGED"
    else:
        remark = "UNDERCHARGED"

    breakdown = (
        f"Courier: SafeExpress | Mode: {mode.title()} | "
        f"Origin: {origin} ({origin_zone}) | Destination: {dest} ({dest_zone}) | "
        f"Matrix Mode: {mode_letter or '-'} | Actual Weight: {actual_weight:.2f} | "
        f"Chargeable Weight: {chargeable_weight:.2f} | Rate: {rate:.2f} | "
        f"Freight: {freight:.2f} | Fuel: {fuel:.2f} | Waybill Charge: {waybill_charge:.2f} | "
        f"UCC: {ucc:.2f} | Area Charge: {area_charge:.2f} | GST: {gst:.2f} |"
        f"Value Surcharge: {value_surcharge:.2f} | "  
    )

    return {
        "Docket No": waybill,
        "Origin": origin,
        "Destination": dest,
        "Zone Origin": origin_zone,
        "Zone Dest": dest_zone,
        "Mode": mode.title(),
        "Weight KG": round(actual_weight, 2),
        "Invoice Amount": round(invoice_amount, 2),
        "Correct Amount": round(correct_total, 2),
        "Difference": round(diff, 2),
        "Remark": remark,
        "Breakdown": breakdown,
    }

# =========================================================
# UNIFIED RUNNER
# =========================================================
def run_audit(courier: str, mode: str, pdf_bytes: bytes):
    courier = (courier or "").strip().lower()
    mode = (mode or "").strip().lower()

    if courier == "dtdc":
        if mode != "surface":
            return [], "DTDC Air not configured yet."
        rows = parse_dtdc_rows(pdf_bytes)
        return [audit_dtdc_surface(r) for r in rows], ""

    if courier == "safeexpress":
        rows = parse_safeexpress_rows(pdf_bytes)
        return [calculate_safeexpress(r, mode) for r in rows], ""

    return [], "Unsupported courier."

# =========================================================
# HTML
# =========================================================
HTML = """
<!doctype html>
<html>
<head>
  <meta charset="utf-8">
  <title>DTDC Invoice Audit Tool</title>
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <style>
    body{font-family:Arial,sans-serif;background:#f4f6fb;margin:0;padding:20px;color:#111827}
    .card{max-width:1350px;margin:0 auto;background:#fff;border:1px solid #ddd;border-radius:14px;padding:18px;box-shadow:0 10px 25px rgba(0,0,0,.05)}
    h1{margin:0 0 8px;font-size:28px}
    .sub{color:#6b7280;font-size:14px;margin-bottom:14px}
    .row{display:flex;gap:12px;flex-wrap:wrap;align-items:flex-end}
    .col{flex:1;min-width:240px}
    label{display:block;font-size:12px;color:#666;margin-bottom:6px}
    select,input[type=file],button{width:100%;padding:10px;border-radius:10px;border:1px solid #cbd5e1}
    button{background:#0f172a;color:#fff;cursor:pointer;border:none;font-weight:600}
    button:disabled{background:#94a3b8;cursor:not-allowed}
    .hint{font-size:12px;color:#6b7280;margin-top:6px}
    .error{background:#fee2e2;border:1px solid #ef4444;color:#991b1b;padding:10px;border-radius:10px;margin:12px 0}
    .info{background:#eff6ff;border:1px solid #93c5fd;color:#1d4ed8;padding:10px;border-radius:10px;margin:12px 0}
    table{width:100%;border-collapse:collapse;margin-top:14px;font-size:12px}
    th,td{border:1px solid #e5e7eb;padding:8px;text-align:left;vertical-align:top}
    th{background:#f8fafc}
    .viewbtn{padding:6px 10px;border:none;border-radius:8px;background:#1d4ed8;color:#fff;cursor:pointer}
    .detail-row{display:none;background:#f8fafc}
    .detail-box{padding:10px;line-height:1.6;white-space:normal}
    .dl{display:inline-block;margin-top:12px;text-decoration:none;color:#1d4ed8}
  </style>
</head>
<body>
  <div class="card">
    <h1>DTDC Invoice Audit Tool</h1>
   <!--<div class="sub"> • 2 couriers • DTDC Surface + SafeExpress Surface/Air</div>-->

    {% if dataset_status %}
      <div class="info">{{ dataset_status }}</div>
    {% endif %}
    {% if error %}
      <div class="error">{{ error }}</div>
    {% endif %}
    {% if info %}
      <div class="info">{{ info }}</div>
    {% endif %}

    <form action="/audit" method="post" enctype="multipart/form-data">
      <div class="row">
        <div class="col">
          <label>Courier</label>
          <select name="courier" id="courier">
            <option value="">-- Select Courier --</option>
            <option value="dtdc" {% if courier == 'dtdc' %}selected{% endif %}>DTDC</option>
          </select>
        </div>

        <div class="col">
          <label>Mode</label>
          <select name="mode" id="mode">
            <option value="">-- Select Mode --</option>
            <option value="surface" {% if mode == 'surface' %}selected{% endif %}>Surface</option>
            <option value="air" {% if mode == 'air' %}selected{% endif %}>Air</option>
          </select>
          <div class="hint">DTDC Air placeholder.</div>
        </div>

        <div class="col">
          <label>Upload Invoice PDF</label>
          <input type="file" name="pdf" id="pdf" accept="application/pdf">
        </div>

        <div class="col">
          <label>&nbsp;</label>
          <button type="submit" id="runBtn"> Run Audit </button>
        </div>
      </div>
    </form>

    {% if rows and rows|length > 0 %}
      <a class="dl" href="/download">Download Excel</a>
      <table>
        <thead>
          <tr>
            <th>Docket / Waybill</th>
            <th>Origin</th>
            <th>Destination</th>
            <th>Origin Zone</th>
            <th>Destination Zone</th>
            <th>Mode</th>
            <th>Weight KG</th>
            <th>Invoice Amount</th>
            <th>Correct Amount</th>
            <th>Difference</th>
            <th>Remark</th>
            <th>View</th>
          </tr>
        </thead>
        <tbody>
          {% for r in rows %}
            <tr>
              <td>{{ r["Docket No"] }}</td>
              <td>{{ r["Origin"] }}</td>
              <td>{{ r["Destination"] }}</td>
              <td>{{ r["Zone Origin"] }}</td>
              <td>{{ r["Zone Dest"] }}</td>
              <td>{{ r["Mode"] }}</td>
              <td>{{ r["Weight KG"] }}</td>
              <td>{{ r["Invoice Amount"] }}</td>
              <td>{{ r["Correct Amount"] }}</td>
              <td>{{ r["Difference"] }}</td>
              <td>{{ r["Remark"] }}</td>
              <td><button type="button" class="viewbtn" onclick="toggleRow('d{{ loop.index }}')">View</button></td>
            </tr>
            <tr id="d{{ loop.index }}" class="detail-row">
              <td colspan="12">
                <div class="detail-box">{{ r["Breakdown"] }}</div>
              </td>
            </tr>
          {% endfor %}
        </tbody>
      </table>
    {% endif %}
  </div>

<script>
  const courier = document.getElementById("courier");
  const mode = document.getElementById("mode");
  const pdf = document.getElementById("pdf");
  const runBtn = document.getElementById("runBtn");

  function syncUI(){
    const c = courier.value;
    const m = mode.value;
    const enabled = !!c && !!m;
    pdf.disabled = !enabled;
    runBtn.disabled = !enabled;
  }

  function toggleRow(id){
    const row = document.getElementById(id);
    if(!row) return;
    row.style.display = row.style.display === "table-row" ? "none" : "table-row";
  }

  courier.addEventListener("change", syncUI);
  mode.addEventListener("change", syncUI);
  window.addEventListener("load", syncUI);
</script>
</body>
</html>
"""

# =========================================================
# ROUTES
# =========================================================
@app.route("/", methods=["GET"])
def home():
    load_india_cities_csv()
    return render_template_string(
        HTML,
        dataset_status=CITY_DATASET_STATUS,
        error="",
        info="",
        rows=[],
        courier="",
        mode=""
    )


@app.route("/audit", methods=["POST"])
def audit():
    global LAST_ROWS
    load_india_cities_csv()

    courier = (request.form.get("courier") or "").strip().lower()
    mode = (request.form.get("mode") or "").strip().lower()
    pdf_file = request.files.get("pdf")

    if courier not in ("dtdc",):
        return render_template_string(
            HTML, dataset_status=CITY_DATASET_STATUS, error="Please select Courier.",
            info="", rows=[], courier=courier, mode=mode
        )

    if mode not in ("surface", "air"):
        return render_template_string(
            HTML, dataset_status=CITY_DATASET_STATUS, error="Please select Mode.",
            info="", rows=[], courier=courier, mode=mode
        )

    if not pdf_file:
        return render_template_string(
            HTML, dataset_status=CITY_DATASET_STATUS, error="Please upload PDF.",
            info="", rows=[], courier=courier, mode=mode
        )

    pdf_bytes = pdf_file.read()
    rows, err = run_audit(courier, mode, pdf_bytes)

    if err:
        return render_template_string(
            HTML, dataset_status=CITY_DATASET_STATUS, error=err,
            info="", rows=[], courier=courier, mode=mode
        )

    if not rows:
        try:
            ocr_text = ocr_extract_text(pdf_bytes, max_pages=3)
            return render_template_string(
                HTML,
                dataset_status=CITY_DATASET_STATUS,
                error="No billing rows detected in text layer.",
                info=ocr_text[:1500],
                rows=[],
                courier=courier,
                mode=mode
            )
        except Exception as e:
            return render_template_string(
                HTML,
                dataset_status=CITY_DATASET_STATUS,
                error=f"No billing rows detected. OCR failed: {e}",
                info="",
                rows=[],
                courier=courier,
                mode=mode
            )

    LAST_ROWS = rows

    return render_template_string(
        HTML,
        dataset_status=CITY_DATASET_STATUS,
        error="",
        info="",
        rows=rows,
        courier=courier,
        mode=mode
    )


@app.route("/download", methods=["GET"])
def download():
    global LAST_ROWS
    if not LAST_ROWS:
        return "No data to download. Run audit first.", 400

    df = pd.DataFrame(LAST_ROWS)

    out = io.BytesIO()

    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Audit")
        ws = writer.sheets["Audit"]

        # Header style
        header_fill = PatternFill(fill_type="solid", start_color="D9EAF7", end_color="D9EAF7")
        ok_fill = PatternFill(fill_type="solid", start_color="C6EFCE", end_color="C6EFCE")
        over_fill = PatternFill(fill_type="solid", start_color="FFC7CE", end_color="FFC7CE")
        under_fill = PatternFill(fill_type="solid", start_color="FFEB9C", end_color="FFEB9C")

        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.fill = header_fill

        # Find Remark column
        remark_col = None
        for cell in ws[1]:
            if str(cell.value).strip().lower() in ["remark", "remarks"]:
                remark_col = cell.column
                break

        # All cells alignment + row color by remark
        for row in ws.iter_rows(min_row=2):
            remark_value = ""
            if remark_col:
                remark_value = str(row[remark_col - 1].value).strip().upper()

            fill_to_apply = None
            if remark_value == "OK":
                fill_to_apply = ok_fill
            elif remark_value == "OVERCHARGED":
                fill_to_apply = over_fill
            elif remark_value == "UNDERCHARGED":
                fill_to_apply = under_fill

            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                if fill_to_apply:
                    cell.fill = fill_to_apply

        # Auto column width
        for column_cells in ws.columns:
            max_length = 0
            column_letter = column_cells[0].column_letter

            for cell in column_cells:
                try:
                    cell_value = "" if cell.value is None else str(cell.value)
                    if len(cell_value) > max_length:
                        max_length = len(cell_value)
                except:
                    pass

            adjusted_width = min(max_length + 4, 45)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Freeze header row
        ws.freeze_panes = "A2"

        # Auto filter
        ws.auto_filter.ref = ws.dimensions

    out.seek(0)

    return send_file(
        out,
        as_attachment=True,
        download_name=f"Audit_Output_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


if __name__ == "__main__":
    load_india_cities_csv()
    app.run(debug=True)